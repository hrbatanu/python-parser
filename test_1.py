#!/usr/bin/env python
# -*- coding: iso-8859-15 -*-
 
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
import requests
import os
import sys

class Request(object):
    """docstring for Request"""
    def __init__(self, url):
        super(Request, self).__init__()#super returns a proxy object that delegates method calls to a parent or sibling class of type. This is useful for accessing inherited methods that have been overridden in a class. The search order is same as that used by getattr() except that the type itself is skipped.”

        self.url = url
        self.headers = {'Content-Type': 'application/json'} #what is content?

    def make(self, query=None, resource=None):
        if query:
            params = {'qt': 'worldcat_org_all'}
            params['q'] = query
            url = self.url + '/search'#?
        elif resource:
            url = self.url + resource
            params = None
        r = requests.get(url, params=params, headers=self.headers)
        if r.status_code == requests.codes.ok:##where to find?
            return r.text
        else:
            try:
                r.raise_for_status()
            except Exception, e:
                if r.status_code == 403:
                    print "\n\n===================   Execution stopped!   ==================="
                    sys.exit(e)
            return None

class ResponseManager(object):
    def __init__(self):
        super(ResponseManager, self).__init__()

    def get_resource(self, html_doc, text_to_find):
        """This obtains:
        1. view all editions
        2. next pages
        """
        soup = BeautifulSoup(html_doc, 'html.parser')
        links = soup.find_all('a')
        resource = None
        for link in links:
            if text_to_find in link.text.lower() and not resource:
                resource = link.get('href')
        return resource

    def get_resource_titles(self, html_doc):
        """
        This method returns all resources related to titles.
        """
        soup = BeautifulSoup(html_doc, 'html.parser')
        links = soup.find_all('a')
        resources = []
        for link in links:
            href = link.get('href') #get id a dict method returns a value for the given key
            if href and '/title' in href and not href in resources:
                resources.append(href)
        return resources

    def get_ISBN_code(self, html_doc):
        soup = BeautifulSoup(html_doc, 'html.parser')
        tr_element = soup.find(id="details-standardno")
        if tr_element:
            return tr_element.td.string
        else:
            return None


def get_resource_titles_all_pages(html_doc, resources, r_manager):
    resource = r_manager.get_resource(html_doc, 'next')
    if resource:
        os.system('cls' if os.name == 'nt' else 'clear') #Execute the command (a string) in a subshell
        print "getting...", resource
        html_doc = request.make(resource=resource)
        resources_tmp = r_manager.get_resource_titles(html_doc)#returns all resources related to titles
        resources += resources_tmp
        return get_resource_titles_all_pages(html_doc, resources, r_manager)
    else:
        return resources


class WorkbookManager(object):
    """docstring for WorkbookManager"""
    def __init__(self, filename, mode_create=False):
        super(WorkbookManager, self).__init__()
        self.filename = filename
        if mode_create:
            self.wb = Workbook()
            self.ws1 = self.wb.active
            self.ws1.title = "out_put"
            self.current_ISBN_code = None
            self.current_column = 1
            self.current_row = 1
            self.ws1.cell(column=self.current_column, row=self.current_row, value="Original ISBN")

    def get_ISBN_codes(self):#get all 10-digit ISBN from the input csv file.
        try:
            wb = load_workbook(filename = self.filename)
        except Exception, e:
            sys.exit(e)
        sheet_ranges = wb['Sheet1']
        #print sheet_ranges
        codes = []
        repeated_codes = []
        invalid_codes = []
        for s_range in sheet_ranges:
            #print s_range
            ISBN_code = str(s_range[0].value) #just the first column (10-digit)
            if not ISBN_code in codes:
 #               if len(ISBN_code) == 10:
                 codes.append(ISBN_code)
                #else:
 #                   invalid_codes.append(ISBN_code)
            else:
                if not ISBN_code in repeated_codes:
                    repeated_codes.append(ISBN_code)
        return codes
    def get_ISBN_codes13(self):#get all 13-digit ISBN from the input csv file.
        try:
            wb = load_workbook(filename = self.filename)
        except Exception, e:
            sys.exit(e)
        sheet_ranges = wb['Sheet1']
        codes13 = []
        repeated_codes13 = []
        invalid_codes13 = []
        for s_range in sheet_ranges:
            #print s_range
       #just the first column (10-digit)
            ISBN_code13 = str(s_range[1].value)
            if not ISBN_code13 in codes13:
                if len(ISBN_code13) == 13:
                    codes13.append(ISBN_code13)
                else:
                    invalid_codes13.append(ISBN_code13)
            else:
                if not ISBN_code13 in repeated_codes13:
                    repeated_codes13.append(ISBN_code13)
        return codes13            

    def insert_ISBN_source(self, ISBN_code):
        self.current_row += 1
        self.current_column = 1
        print "inserting ISBN code", ISBN_code
        self.ws1.cell(column=self.current_column, row=self.current_row, value=ISBN_code)
    def insert_ISBN_source13(self, ISBN_code13):
        self.current_row += 1
        self.current_column = 2
        print "inserting ISBN code13", ISBN_code13
        self.ws1.cell(column=self.current_column, row=self.current_row, value=ISBN_code13)
    def insert_ISBN_related(self, ISBN_code, ISBN_code_related):
        #find related pair of ISBNs
        if ISBN_code != self.current_ISBN_code:
            self.current_ISBN_code = ISBN_code
            self.current_column = 3 #this is the insertion point of the output
        else:
            self.current_column += 1
        os.system('cls' if os.name == 'nt' else 'clear')
        print "inserting (", ISBN_code_related, ") related to ", ISBN_code
        self.ws1.cell(column=self.current_column, row=self.current_row, value=ISBN_code_related)

    def save(self):
        self.wb.save(filename=self.filename)


wb_manager = WorkbookManager('Workbook1.xlsx')
wb_manager_output = WorkbookManager('new_isbn copy.xlsx', mode_create=True)
ISBN_codes = wb_manager.get_ISBN_codes()
ISBN_codes13 = wb_manager.get_ISBN_codes13()


r_manager = ResponseManager()
request = Request('https://www.worldcat.org')

codes_not_found = []
for ISBN_code in ISBN_codes:
    html_doc = request.make(query=ISBN_code)#search this ISBN
    resource = r_manager.get_resource(html_doc, 'view all editions')
    if resource:
        wb_manager_output.insert_ISBN_source(ISBN_code)#insert original 10-ISBN
        html_doc = request.make(resource=resource)
        resources = r_manager.get_resource_titles(html_doc)
        resources = get_resource_titles_all_pages(html_doc, resources, r_manager)
        for resource in resources:
            html_doc = request.make(resource=resource)
            if html_doc:
                ISBN_code_related = r_manager.get_ISBN_code(html_doc)
                if ISBN_code_related:
                    wb_manager_output.insert_ISBN_related(ISBN_code, ISBN_code_related)
                    wb_manager_output.save()
    else:
        codes_not_found.append(ISBN_code)
#create a paralle one!
for ISBN_code13 in ISBN_codes13:
    html_doc = request.make(query=ISBN_code13)#search this ISBN-13
    resource = r_manager.get_resource(html_doc, 'view all editions')
    if resource:
        wb_manager_output.insert_ISBN_source13(ISBN_code13)#insert original 13-ISBN
#        html_doc = request.make(resource=resource)
#        resources = r_manager.get_resource_titles(html_doc)
 #       resources = get_resource_titles_all_pages(html_doc, resources, r_manager)
        #for resource in resources:
            #html_doc = request.make(resource=resource)
            #if html_doc:
    #            ISBN_code_related13 = r_manager.get_ISBN_code(html_doc)
 #               if ISBN_code_related13:
 #                   wb_manager_output.insert_ISBN_related13(ISBN_code13, ISBN_code_related13)
#                    wb_manager_output.save()###
    else:
        codes_not_found.append(ISBN_code13)

wb_manager_output.save()

print "Finished!"

